from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count, F
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
from .models import (
    Product, Category, Supplier, Location, StockMovement, Alert,
    Customer, Sale, SaleItem, InventoryTransfer, TransferItem,
    CustomerSaleHistory, CustomerNote, CustomerAlert, InvoiceTemplate,
    LocationAlert, ProductLocationStock, SystemSettings
)
from .forms import (
    ProductForm, CategoryForm, SupplierForm, LocationForm,
    StockMovementForm, AlertForm, ProductSearchForm, BulkUpdateForm,
    CustomerForm, SaleForm, SaleItemForm, InventoryTransferForm, TransferItemForm,
    SaleItemFormSet, TransferItemFormSet, CustomerNoteForm, CustomerAlertForm,
    InvoiceTemplateForm, CustomerSearchForm, SystemSettingsForm
)
import barcode
from barcode.writer import ImageWriter
from io import BytesIO
import base64
import csv
import json
from datetime import timedelta
import json
from datetime import datetime, timedelta
import io

def dashboard(request):
    """דשבורד ראשי עם סטטיסטיקות"""
    # סטטיסטיקות כלליות (מכל המיקומים)
    total_products = Product.objects.count()
    low_stock_products = Product.objects.filter(quantity__lte=F('min_quantity')).count()
    out_of_stock_products = Product.objects.filter(quantity=0).count()
    total_stock_value = Product.objects.aggregate(
        total=Sum(F('quantity') * F('cost_price'))
    )['total'] or 0

    # סטטיסטיקות מחסן (מתוך ProductLocationStock)
    warehouse_locations = Location.objects.filter(location_type='warehouse', is_active=True)
    warehouse_stock = ProductLocationStock.objects.filter(location__in=warehouse_locations)
    warehouse_total = warehouse_stock.values('product').distinct().count()
    warehouse_quantity = warehouse_stock.aggregate(total=Sum('quantity'))['total'] or 0
    warehouse_low_stock = warehouse_stock.filter(quantity__lte=F('min_quantity')).count()
    warehouse_out_of_stock = warehouse_stock.filter(quantity=0).count()
    warehouse_value = warehouse_stock.aggregate(
        total=Sum(F('quantity') * F('product__cost_price'))
    )['total'] or 0

    # סטטיסטיקות חנות (מתוך ProductLocationStock)
    store_locations = Location.objects.filter(location_type='store', is_active=True)
    store_stock = ProductLocationStock.objects.filter(location__in=store_locations)
    store_total = store_stock.values('product').distinct().count()
    store_quantity = store_stock.aggregate(total=Sum('quantity'))['total'] or 0
    store_low_stock = store_stock.filter(quantity__lte=F('min_quantity')).count()
    store_out_of_stock = store_stock.filter(quantity=0).count()
    store_value = store_stock.aggregate(
        total=Sum(F('quantity') * F('product__cost_price'))
    )['total'] or 0

    # התראות
    alerts = Alert.objects.filter(is_resolved=False).order_by('-created_at')[:10]

    # התראות מיקום
    location_alerts = LocationAlert.objects.filter(is_resolved=False).order_by('-created_at')[:10]

    # מוצרים עם מלאי נמוך במחסן
    warehouse_low_stock_items = warehouse_stock.filter(quantity__lte=F('min_quantity')).select_related('product').order_by('quantity')[:5]

    # מוצרים עם מלאי נמוך בחנות
    store_low_stock_items = store_stock.filter(quantity__lte=F('min_quantity')).select_related('product').order_by('quantity')[:5]

    # תנועות מלאי אחרונות
    recent_movements = StockMovement.objects.select_related('product').order_by('-created_at')[:10]

    # סטטיסטיקות לפי קטגוריה
    category_stats = Product.objects.values('category__name').annotate(
        count=Count('id'),
        total_value=Sum(F('quantity') * F('cost_price'))
    ).order_by('-count')[:5]

    # סטטיסטיקות לפי מיקום
    location_stats = Product.objects.values('location__name', 'location__location_type').annotate(
        count=Count('id'),
        total_quantity=Sum('quantity'),
        total_value=Sum(F('quantity') * F('cost_price'))
    ).order_by('-total_value')[:10]

    # לקוחות אחרונים
    recent_customers = Customer.objects.order_by('-created_at')[:5]

    # מכירות אחרונות
    recent_sales = Sale.objects.select_related('customer').order_by('-created_at')[:5]

    context = {
        'total_products': total_products,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'total_stock_value': total_stock_value,

        # מחסן
        'warehouse_total': warehouse_total,
        'warehouse_quantity': warehouse_quantity,
        'warehouse_low_stock': warehouse_low_stock,
        'warehouse_out_of_stock': warehouse_out_of_stock,
        'warehouse_value': warehouse_value,
        'warehouse_low_stock_items': warehouse_low_stock_items,

        # חנות
        'store_total': store_total,
        'store_quantity': store_quantity,
        'store_low_stock': store_low_stock,
        'store_out_of_stock': store_out_of_stock,
        'store_value': store_value,
        'store_low_stock_items': store_low_stock_items,

        'alerts': alerts,
        'location_alerts': location_alerts,
        'recent_movements': recent_movements,
        'category_stats': category_stats,
        'location_stats': location_stats,
        'recent_customers': recent_customers,
        'recent_sales': recent_sales,
    }
    return render(request, 'inventory/dashboard.html', context)

def product_list(request):
    """רשימת מוצרים עם חיפוש וסינון"""
    search_form = ProductSearchForm(request.GET)
    products = Product.objects.select_related('category', 'supplier', 'location')

    if search_form.is_valid():
        search = search_form.cleaned_data.get('search')
        category = search_form.cleaned_data.get('category')
        supplier = search_form.cleaned_data.get('supplier')
        status = search_form.cleaned_data.get('status')
        low_stock = search_form.cleaned_data.get('low_stock')
        out_of_stock = search_form.cleaned_data.get('out_of_stock')

        if search:
            products = products.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(barcode__icontains=search) |
                Q(description__icontains=search)
            )

        if category:
            products = products.filter(category=category)

        if supplier:
            products = products.filter(supplier=supplier)

        if status:
            products = products.filter(status=status)

        if low_stock:
            products = products.filter(quantity__lte=F('min_quantity'))

        if out_of_stock:
            products = products.filter(quantity=0)

    # מיון
    sort_by = request.GET.get('sort', 'name')
    if sort_by == 'quantity':
        products = products.order_by('quantity')
    elif sort_by == 'value':
        products = products.order_by(F('quantity') * F('cost_price'))
    elif sort_by == 'updated':
        products = products.order_by('-updated_at')
    else:
        products = products.order_by('name')

    # עמודים
    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    products = paginator.get_page(page_number)

    context = {
        'products': products,
        'search_form': search_form,
        'sort_by': sort_by,
    }
    return render(request, 'inventory/product_list.html', context)

def add_product(request):
    """הוספת מוצר חדש"""
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                product = form.save(commit=False)
                # הגדרת ערכי ברירת מחדל לשדות חובה
                if not product.sku:
                    product.sku = f"SKU-{Product.objects.count() + 1:06d}"
                if not product.status:
                    product.status = 'active'
                if not product.unit:
                    product.unit = 'pcs'
                if not product.cost_price:
                    product.cost_price = 0
                if not product.selling_price:
                    product.selling_price = 0
                if not product.min_quantity:
                    product.min_quantity = 0
                if not product.max_quantity:
                    product.max_quantity = 1000

                product.save()
                messages.success(request, 'המוצר נוסף בהצלחה!')
                return redirect('product_detail', pk=product.pk)
            except Exception as e:
                messages.error(request, f'שגיאה בשמירת המוצר: {str(e)}')
        else:
            messages.error(request, 'יש שגיאות בטופס. אנא בדוק את הנתונים.')
    else:
        form = ProductForm()
    return render(request, 'inventory/add_product.html', {'form': form})

# @login_required
def product_detail(request, pk):
    """פרטי מוצר"""
    product = get_object_or_404(Product, pk=pk)
    movements = StockMovement.objects.filter(product=product).order_by('-created_at')[:20]
    alerts = Alert.objects.filter(product=product, is_resolved=False)

    context = {
        'product': product,
        'movements': movements,
        'alerts': alerts,
    }
    return render(request, 'inventory/product_detail.html', context)

# @login_required
def edit_product(request, pk):
    """עריכת מוצר"""
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'המוצר עודכן בהצלחה!')
            return redirect('product_detail', pk=product.pk)
    else:
        form = ProductForm(instance=product)
    return render(request, 'inventory/edit_product.html', {'form': form, 'product': product})

def delete_product(request, pk):
    """מחיקת מוצר"""
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'POST':
        product_name = product.name
        try:
            # בדיקה אם יש מכירות קשורות
            sales_count = SaleItem.objects.filter(product=product).count()
            if sales_count > 0:
                messages.error(request, f'לא ניתן למחוק את המוצר "{product_name}" כי יש לו {sales_count} מכירות קשורות. שנה את הסטטוס ל"הופסק" במקום.')
                return redirect('product_detail', pk=pk)

            # בדיקה אם יש תנועות מלאי
            movements_count = StockMovement.objects.filter(product=product).count()

            # מחיקת המוצר (ימחק גם ProductLocationStock אוטומטית בגלל CASCADE)
            product.delete()
            messages.success(request, f'המוצר "{product_name}" נמחק בהצלחה!')
            return redirect('product_list')
        except Exception as e:
            messages.error(request, f'שגיאה במחיקת המוצר: {str(e)}')
            return redirect('product_detail', pk=pk)

    # בדיקה אם יש רשומות קשורות
    sales_count = SaleItem.objects.filter(product=product).count()
    movements_count = StockMovement.objects.filter(product=product).count()
    location_stocks = ProductLocationStock.objects.filter(product=product).count()

    return render(request, 'inventory/delete_product.html', {
        'product': product,
        'sales_count': sales_count,
        'movements_count': movements_count,
        'location_stocks': location_stocks,
        'can_delete': sales_count == 0
    })

# @login_required
def stock_movement(request, pk):
    """תנועת מלאי"""
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        form = StockMovementForm(request.POST, product=product)
        if form.is_valid():
            with transaction.atomic():
                movement = form.save(commit=False)
                movement.product = product
                # movement.user = request.user  # זמנית ללא משתמש
                movement.previous_quantity = product.quantity

                # עדכון כמות המלאי
                if movement.movement_type == 'in':
                    product.quantity += movement.quantity
                elif movement.movement_type == 'out':
                    if product.quantity >= movement.quantity:
                        product.quantity -= movement.quantity
                    else:
                        messages.error(request, 'אין מספיק מלאי!')
                        return render(request, 'inventory/stock_movement.html', {'form': form, 'product': product})
                elif movement.movement_type == 'adjustment':
                    product.quantity = movement.quantity

                movement.new_quantity = product.quantity
                product.save()
                movement.save()

                # יצירת התראות
                create_stock_alerts(product)

                messages.success(request, 'תנועת המלאי בוצעה בהצלחה!')
                return redirect('product_detail', pk=product.pk)
    else:
        form = StockMovementForm(product=product)
    return render(request, 'inventory/stock_movement.html', {'form': form, 'product': product})

def create_stock_alerts(product):
    """יצירת התראות מלאי"""
    if product.quantity <= 0:
        Alert.objects.create(
            product=product,
            alert_type='out_of_stock',
            priority='critical',
            message=f'המוצר {product.name} אזל מהמלאי!'
        )
    elif product.quantity <= product.min_quantity:
        Alert.objects.create(
            product=product,
            alert_type='low_stock',
            priority='high',
            message=f'המוצר {product.name} במלאי נמוך ({product.quantity} יחידות)'
        )
    elif product.quantity >= product.max_quantity:
        Alert.objects.create(
            product=product,
            alert_type='overstock',
            priority='medium',
            message=f'המוצר {product.name} במלאי עודף ({product.quantity} יחידות)'
        )

# @login_required
def bulk_update(request):
    """עדכון מרובה"""
    try:
        products = Product.objects.all()[:50]  # הגבלה ל-50 מוצרים לתצוגה
    except Exception as e:
        products = []
        print(f"Error fetching products: {e}")

    if request.method == 'POST':
        form = BulkUpdateForm(request.POST)
        product_ids = request.POST.getlist('product_ids')

        if form.is_valid() and product_ids:
            action = form.cleaned_data['action']
            value = form.cleaned_data['value']
            reason = form.cleaned_data['reason']

            products_to_update = Product.objects.filter(id__in=product_ids)
            updated_count = 0

            try:
                with transaction.atomic():
                    for product in products_to_update:
                        old_quantity = product.quantity

                        if action == 'add':
                            product.quantity += int(value or 0)
                        elif action == 'subtract':
                            product.quantity = max(0, product.quantity - int(value or 0))
                        elif action == 'set':
                            product.quantity = int(value or 0)
                        elif action == 'update_price':
                            product.selling_price = value or product.selling_price
                        elif action == 'change_status':
                            product.status = form.cleaned_data['new_status']
                        elif action == 'change_category':
                            product.category = form.cleaned_data['new_category']

                        product.save()

                        # יצירת תנועת מלאי
                        if action in ['add', 'subtract', 'set']:
                            StockMovement.objects.create(
                                product=product,
                                movement_type='adjustment',
                                quantity=product.quantity - old_quantity,
                                previous_quantity=old_quantity,
                                new_quantity=product.quantity,
                                reason=f'עדכון מרובה: {reason}',
                                # user=request.user  # זמנית ללא משתמש
                            )

                        updated_count += 1

                messages.success(request, f'{updated_count} מוצרים עודכנו בהצלחה!')
                return redirect('product_list')

            except Exception as e:
                messages.error(request, f'שגיאה בעדכון: {str(e)}')
        else:
            if not product_ids:
                messages.error(request, 'אנא בחר לפחות מוצר אחד')
            else:
                messages.error(request, 'יש שגיאות בטופס')
    else:
        form = BulkUpdateForm()

    return render(request, 'inventory/bulk_update.html', {'form': form, 'products': products})

# @login_required
def alerts(request):
    """ניהול התראות"""
    alerts = Alert.objects.filter(is_resolved=False).order_by('-created_at')

    if request.method == 'POST':
        alert_id = request.POST.get('alert_id')
        action = request.POST.get('action')

        if alert_id and action:
            alert = get_object_or_404(Alert, id=alert_id)
            if action == 'mark_read':
                alert.is_read = True
                alert.save()
            elif action == 'resolve':
                alert.is_resolved = True
                alert.resolved_at = timezone.now()
                # alert.resolved_by = request.user  # זמנית ללא משתמש
                alert.save()

    return render(request, 'inventory/alerts.html', {'alerts': alerts})

# @login_required
def reports(request):
    """דוחות ואנליטיקה"""
    # סטטיסטיקות כלליות
    total_products = Product.objects.count()
    low_stock_products = Product.objects.filter(quantity__lte=F('min_quantity')).count()
    out_of_stock_products = Product.objects.filter(quantity=0).count()
    total_stock_value = Product.objects.aggregate(
        total=Sum(F('quantity') * F('cost_price'))
    )['total'] or 0

    # דוח מלאי נמוך
    low_stock = Product.objects.filter(quantity__lte=F('min_quantity')).order_by('quantity')

    # דוח ערך מלאי
    stock_value = Product.objects.aggregate(
        total_cost=Sum(F('quantity') * F('cost_price')),
        total_selling=Sum(F('quantity') * F('selling_price'))
    )

    # דוח תנועות מלאי
    movements_today = StockMovement.objects.filter(
        created_at__date=timezone.now().date()
    ).count()

    movements_week = StockMovement.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=7)
    ).count()

    # דוח לפי קטגוריה
    category_report = Product.objects.values('category__name').annotate(
        count=Count('id'),
        total_quantity=Sum('quantity'),
        total_value=Sum(F('quantity') * F('cost_price'))
    ).order_by('-total_value')

    # דוח ספקים
    supplier_stats = Product.objects.values('supplier__name').annotate(
        count=Count('id'),
        total_value=Sum(F('quantity') * F('cost_price'))
    ).order_by('-count')[:10]

    # דוח מוצרים עם הרווח הגבוה ביותר
    high_margin_products = Product.objects.filter(
        margin_percentage__gt=0
    ).order_by('-margin_percentage')[:10]

    # דוח תנועות לפי סוג
    movement_stats = StockMovement.objects.values('movement_type').annotate(
        count=Count('id'),
        total_quantity=Sum('quantity')
    ).order_by('-count')

    # דוח התראות
    active_alerts = Alert.objects.filter(is_resolved=False).count()
    critical_alerts = Alert.objects.filter(is_resolved=False, priority='critical').count()

    context = {
        'total_products': total_products,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'total_stock_value': total_stock_value,
        'low_stock': low_stock,
        'stock_value': stock_value,
        'movements_today': movements_today,
        'movements_week': movements_week,
        'category_report': category_report,
        'supplier_stats': supplier_stats,
        'high_margin_products': high_margin_products,
        'movement_stats': movement_stats,
        'active_alerts': active_alerts,
        'critical_alerts': critical_alerts,
    }
    return render(request, 'inventory/reports.html', context)

def generate_barcode(request):
    """יצירת ברקוד אוטומטי"""
    if request.method == 'POST':
        barcode_text = request.POST.get('barcode_text', '')
        if barcode_text:
            try:
                # יצירת ברקוד EAN13
                ean = barcode.get('ean13', barcode_text, writer=ImageWriter())
                buffer = BytesIO()
                ean.write(buffer)
                buffer.seek(0)

                # המרה ל-base64
                barcode_data = base64.b64encode(buffer.getvalue()).decode()

                return JsonResponse({
                    'success': True,
                    'barcode_data': f'data:image/png;base64,{barcode_data}',
                    'barcode_text': barcode_text
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'No barcode text provided'})

# ניהול קטגוריות
# @login_required
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'inventory/category_list.html', {'categories': categories})

# @login_required
def add_category(request):
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'הקטגוריה נוספה בהצלחה!')
            return redirect('category_list')
    else:
        form = CategoryForm()
    return render(request, 'inventory/add_category.html', {'form': form})

# ניהול ספקים
# @login_required
def supplier_list(request):
    suppliers = Supplier.objects.all()
    return render(request, 'inventory/supplier_list.html', {'suppliers': suppliers})

# @login_required
def add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'הספק נוסף בהצלחה!')
            return redirect('supplier_list')
    else:
        form = SupplierForm()
    return render(request, 'inventory/add_supplier.html', {'form': form})

def edit_category(request, pk):
    """עריכת קטגוריה"""
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, 'הקטגוריה עודכנה בהצלחה!')
            return redirect('category_list')
    else:
        form = CategoryForm(instance=category)
    return render(request, 'inventory/edit_category.html', {'form': form, 'category': category})

def delete_category(request, pk):
    """מחיקת קטגוריה"""
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f'הקטגוריה "{category_name}" נמחקה בהצלחה!')
        return redirect('category_list')
    return render(request, 'inventory/delete_category.html', {'category': category})

def edit_supplier(request, pk):
    """עריכת ספק"""
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, 'הספק עודכן בהצלחה!')
            return redirect('supplier_list')
    else:
        form = SupplierForm(instance=supplier)
    return render(request, 'inventory/edit_supplier.html', {'form': form, 'supplier': supplier})

def delete_supplier(request, pk):
    """מחיקת ספק"""
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == 'POST':
        supplier_name = supplier.name
        supplier.delete()
        messages.success(request, f'הספק "{supplier_name}" נמחק בהצלחה!')
        return redirect('supplier_list')
    return render(request, 'inventory/delete_supplier.html', {'supplier': supplier})

# פונקציות ייצוא דוחות
def export_products_csv(request):
    """ייצוא רשימת מוצרים ל-CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="products.csv"'

    # הוספת BOM ל-UTF-8 לתמיכה ב-Excel
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow([
        'שם מוצר', 'SKU', 'תיאור', 'קטגוריה', 'ספק', 'מיקום',
        'כמות', 'כמות מינימלית', 'כמות מקסימלית', 'יחידה',
        'מחיר עלות', 'מחיר מכירה', 'אחוז רווח', 'ברקוד',
        'סטטוס', 'משקל', 'מידות', 'הערות', 'תאריך יצירה'
    ])

    products = Product.objects.select_related('category', 'supplier').prefetch_related('location_stocks__location').all()
    for product in products:
        # מיקומי המוצר (מתוך ProductLocationStock)
        locations_str = ', '.join([
            f"{pls.location.name}: {pls.quantity}"
            for pls in product.location_stocks.all()
        ]) if product.location_stocks.exists() else 'אין מיקום'

        writer.writerow([
            product.name,
            product.sku or '',
            product.description or '',
            product.category.name if product.category else '',
            product.supplier.name if product.supplier else '',
            locations_str,
            product.quantity,
            product.min_quantity,
            product.max_quantity,
            product.get_unit_display(),
            product.cost_price or 0,
            product.selling_price or 0,
            product.margin_percentage or 0,
            product.barcode or '',
            product.get_status_display(),
            product.weight or '',
            product.dimensions or '',
            product.notes or '',
            product.created_at.strftime('%d/%m/%Y %H:%M') if product.created_at else ''
        ])

    return response

def export_low_stock_csv(request):
    """ייצוא מוצרים במלאי נמוך ל-CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="low_stock_products.csv"'

    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow([
        'שם מוצר', 'SKU', 'כמות נוכחית', 'כמות מינימלית', 'סטטוס',
        'קטגוריה', 'ספק', 'מחיר עלות', 'ערך מלאי'
    ])

    low_stock_products = Product.objects.filter(quantity__lte=F('min_quantity')).select_related('category', 'supplier')
    for product in low_stock_products:
        writer.writerow([
            product.name,
            product.sku or '',
            product.quantity,
            product.min_quantity,
            'אזל מהמלאי' if product.quantity == 0 else 'מלאי נמוך',
            product.category.name if product.category else '',
            product.supplier.name if product.supplier else '',
            product.cost_price or 0,
            product.stock_value or 0
        ])

    return response

def export_movements_csv(request):
    """ייצוא תנועות מלאי ל-CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="stock_movements.csv"'

    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow([
        'מוצר', 'SKU', 'סוג תנועה', 'כמות', 'כמות קודמת', 'כמות חדשה',
        'סיבה', 'התייחסות', 'תאריך'
    ])

    movements = StockMovement.objects.select_related('product').order_by('-created_at')
    for movement in movements:
        writer.writerow([
            movement.product.name,
            movement.product.sku or '',
            movement.get_movement_type_display(),
            movement.quantity,
            movement.previous_quantity,
            movement.new_quantity,
            movement.reason or '',
            movement.reference or '',
            movement.created_at.strftime('%d/%m/%Y %H:%M')
        ])

    return response

def export_reports_json(request):
    """ייצוא דוחות ל-JSON"""
    # סטטיסטיקות כלליות
    total_products = Product.objects.count()
    low_stock_products = Product.objects.filter(quantity__lte=F('min_quantity')).count()
    out_of_stock_products = Product.objects.filter(quantity=0).count()
    total_stock_value = Product.objects.aggregate(
        total=Sum(F('quantity') * F('cost_price'))
    )['total'] or 0

    # דוח לפי קטגוריות
    category_report = []
    for item in Product.objects.values('category__name').annotate(
        count=Count('id'),
        total_value=Sum(F('quantity') * F('cost_price'))
    ).order_by('-total_value'):
        category_report.append({
            'category_name': item['category__name'],
            'count': item['count'],
            'total_value': float(item['total_value'] or 0)
        })

    # דוח ספקים
    supplier_report = []
    for item in Product.objects.values('supplier__name').annotate(
        count=Count('id'),
        total_value=Sum(F('quantity') * F('cost_price'))
    ).order_by('-count'):
        supplier_report.append({
            'supplier_name': item['supplier__name'],
            'count': item['count'],
            'total_value': float(item['total_value'] or 0)
        })

    data = {
        'export_date': timezone.now().strftime('%d/%m/%Y %H:%M'),
        'summary': {
            'total_products': total_products,
            'low_stock_products': low_stock_products,
            'out_of_stock_products': out_of_stock_products,
            'total_stock_value': float(total_stock_value)
        },
        'category_report': category_report,
        'supplier_report': supplier_report
    }

    response = HttpResponse(
        json.dumps(data, ensure_ascii=False, indent=2),
        content_type='application/json; charset=utf-8'
    )
    response['Content-Disposition'] = 'attachment; filename="inventory_report.json"'
    return response

def export_categories_csv(request):
    """ייצוא קטגוריות ל-CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="categories.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow([
        'שם קטגוריה', 'תיאור', 'כמות מוצרים', 'ערך מלאי', 'תאריך יצירה'
    ])

    categories = Category.objects.annotate(
        product_count=Count('product'),
        total_value=Sum(F('product__quantity') * F('product__cost_price'))
    )

    for category in categories:
        writer.writerow([
            category.name,
            category.description or '',
            category.product_count,
            category.total_value or 0,
            category.created_at.strftime('%d/%m/%Y %H:%M') if category.created_at else ''
        ])

    return response

def export_suppliers_csv(request):
    """ייצוא ספקים ל-CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="suppliers.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow([
        'שם ספק', 'איש קשר', 'אימייל', 'טלפון', 'כתובת',
        'כמות מוצרים', 'ערך מלאי', 'פעיל', 'תאריך יצירה'
    ])

    suppliers = Supplier.objects.annotate(
        product_count=Count('product'),
        total_value=Sum(F('product__quantity') * F('product__cost_price'))
    )

    for supplier in suppliers:
        writer.writerow([
            supplier.name,
            supplier.contact_person or '',
            supplier.email or '',
            supplier.phone or '',
            supplier.address or '',
            supplier.product_count,
            supplier.total_value or 0,
            'כן' if supplier.is_active else 'לא',
            supplier.created_at.strftime('%d/%m/%Y %H:%M') if supplier.created_at else ''
        ])

    return response

def export_customers_csv(request):
    """ייצוא לקוחות ל-CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="customers.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow([
        'שם לקוח', 'קוד לקוח', 'אימייל', 'טלפון', 'כתובת', 'עיר', 'מיקוד',
        'ח.פ/ת.ז', 'כמות הזמנות', 'סה"כ קניות', 'הנחה %', 'פעיל', 'תאריך יצירה'
    ])

    customers = Customer.objects.all()

    for customer in customers:
        writer.writerow([
            customer.name,
            customer.customer_code or '',
            customer.email or '',
            customer.phone or '',
            customer.address or '',
            customer.city or '',
            customer.postal_code or '',
            customer.tax_id or '',
            customer.total_orders or 0,
            customer.total_purchases or 0,
            customer.discount_percent or 0,
            'כן' if customer.is_active else 'לא',
            customer.created_at.strftime('%d/%m/%Y %H:%M') if customer.created_at else ''
        ])

    return response

def export_sales_csv(request):
    """ייצוא מכירות ל-CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="sales.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow([
        'מספר חשבונית', 'לקוח', 'מיקום', 'סטטוס', 'סטטוס תשלום', 'אמצעי תשלום',
        'סכום ביניים', 'מע"מ', 'סה"כ', 'הערות', 'תאריך מכירה'
    ])

    sales = Sale.objects.select_related('customer', 'location').all()

    for sale in sales:
        writer.writerow([
            sale.invoice_number,
            sale.customer.name if sale.customer else '',
            sale.location.name if sale.location else '',
            sale.get_status_display(),
            sale.get_payment_status_display(),
            sale.get_payment_method_display() if sale.payment_method else '',
            sale.subtotal or 0,
            sale.tax_amount or 0,
            sale.total_amount or 0,
            sale.notes or '',
            sale.sale_date.strftime('%d/%m/%Y %H:%M') if sale.sale_date else ''
        ])

    return response

def export_locations_csv(request):
    """ייצוא מיקומים ל-CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="locations.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow([
        'שם מיקום', 'סוג מיקום', 'כתובת', 'עיר', 'טלפון',
        'אחראי', 'כמות מוצרים', 'פעיל', 'תאריך יצירה'
    ])

    locations = Location.objects.annotate(
        product_count=Count('product_stocks')
    )

    for location in locations:
        writer.writerow([
            location.name,
            location.get_location_type_display(),
            location.address or '',
            location.city or '',
            location.phone or '',
            location.manager_name or '',
            location.product_count,
            'כן' if location.is_active else 'לא',
            location.created_at.strftime('%d/%m/%Y %H:%M') if location.created_at else ''
        ])

    return response

def export_alerts_csv(request):
    """ייצוא התראות ל-CSV"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="alerts.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow([
        'מוצר', 'סוג התראה', 'רמת חומרה', 'הודעה', 'נפתר', 'תאריך יצירה', 'תאריך פתרון'
    ])

    alerts = Alert.objects.select_related('product').all()

    for alert in alerts:
        writer.writerow([
            alert.product.name if alert.product else '',
            alert.get_alert_type_display(),
            alert.get_priority_display(),
            alert.message or '',
            'כן' if alert.is_resolved else 'לא',
            alert.created_at.strftime('%d/%m/%Y %H:%M') if alert.created_at else '',
            alert.resolved_at.strftime('%d/%m/%Y %H:%M') if alert.resolved_at else ''
        ])

    return response

def export_all_data_csv(request):
    """ייצוא כל הנתונים לקובץ ZIP עם כל הטבלאות"""
    import zipfile
    from io import BytesIO

    # יצירת buffer לזיכרון
    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        # 1. מוצרים
        products_buffer = BytesIO()
        products_buffer.write('\ufeff'.encode('utf-8'))
        writer = csv.writer(io.TextIOWrapper(products_buffer, encoding='utf-8', newline=''))
        writer.writerow([
            'שם מוצר', 'SKU', 'תיאור', 'קטגוריה', 'ספק', 'מיקום',
            'כמות', 'כמות מינימלית', 'כמות מקסימלית', 'יחידה',
            'מחיר עלות', 'מחיר מכירה', 'אחוז רווח', 'ברקוד',
            'סטטוס', 'משקל', 'מידות', 'הערות', 'תאריך יצירה'
        ])
        products = Product.objects.select_related('category', 'supplier').prefetch_related('location_stocks__location').all()
        for product in products:
            locations_str = ', '.join([
                f"{pls.location.name}: {pls.quantity}"
                for pls in product.location_stocks.all()
            ]) if product.location_stocks.exists() else 'אין מיקום'
            writer.writerow([
                product.name, product.sku or '', product.description or '',
                product.category.name if product.category else '',
                product.supplier.name if product.supplier else '',
                locations_str, product.quantity, product.min_quantity, product.max_quantity,
                product.get_unit_display(), product.cost_price or 0,
                product.selling_price or 0, product.margin_percentage or 0,
                product.barcode or '', product.get_status_display(),
                product.weight or '', product.dimensions or '',
                product.notes or '', product.created_at.strftime('%d/%m/%Y %H:%M') if product.created_at else ''
            ])
        zip_file.writestr('1_products.csv', products_buffer.getvalue())

        # 2. קטגוריות
        categories_buffer = BytesIO()
        categories_buffer.write('\ufeff'.encode('utf-8'))
        writer = csv.writer(io.TextIOWrapper(categories_buffer, encoding='utf-8', newline=''))
        writer.writerow(['שם קטגוריה', 'תיאור', 'כמות מוצרים', 'ערך מלאי', 'תאריך יצירה'])
        categories = Category.objects.annotate(
            product_count=Count('product'),
            total_value=Sum(F('product__quantity') * F('product__cost_price'))
        )
        for category in categories:
            writer.writerow([
                category.name, category.description or '', category.product_count,
                category.total_value or 0,
                category.created_at.strftime('%d/%m/%Y %H:%M') if category.created_at else ''
            ])
        zip_file.writestr('2_categories.csv', categories_buffer.getvalue())

        # 3. ספקים
        suppliers_buffer = BytesIO()
        suppliers_buffer.write('\ufeff'.encode('utf-8'))
        writer = csv.writer(io.TextIOWrapper(suppliers_buffer, encoding='utf-8', newline=''))
        writer.writerow([
            'שם ספק', 'איש קשר', 'אימייל', 'טלפון', 'כתובת',
            'כמות מוצרים', 'ערך מלאי', 'פעיל', 'תאריך יצירה'
        ])
        suppliers = Supplier.objects.annotate(
            product_count=Count('product'),
            total_value=Sum(F('product__quantity') * F('product__cost_price'))
        )
        for supplier in suppliers:
            writer.writerow([
                supplier.name, supplier.contact_person or '', supplier.email or '',
                supplier.phone or '', supplier.address or '', supplier.product_count,
                supplier.total_value or 0, 'כן' if supplier.is_active else 'לא',
                supplier.created_at.strftime('%d/%m/%Y %H:%M') if supplier.created_at else ''
            ])
        zip_file.writestr('3_suppliers.csv', suppliers_buffer.getvalue())

        # 4. לקוחות
        customers_buffer = BytesIO()
        customers_buffer.write('\ufeff'.encode('utf-8'))
        writer = csv.writer(io.TextIOWrapper(customers_buffer, encoding='utf-8', newline=''))
        writer.writerow([
            'שם לקוח', 'קוד לקוח', 'אימייל', 'טלפון', 'כתובת', 'עיר', 'מיקוד',
            'ח.פ/ת.ז', 'כמות הזמנות', 'סה"כ קניות', 'הנחה %', 'פעיל', 'תאריך יצירה'
        ])
        customers = Customer.objects.all()
        for customer in customers:
            writer.writerow([
                customer.name, customer.customer_code or '', customer.email or '', customer.phone or '',
                customer.address or '', customer.city or '', customer.postal_code or '',
                customer.tax_id or '', customer.total_orders or 0,
                customer.total_purchases or 0, customer.discount_percent or 0,
                'כן' if customer.is_active else 'לא',
                customer.created_at.strftime('%d/%m/%Y %H:%M') if customer.created_at else ''
            ])
        zip_file.writestr('4_customers.csv', customers_buffer.getvalue())

        # 5. מכירות
        sales_buffer = BytesIO()
        sales_buffer.write('\ufeff'.encode('utf-8'))
        writer = csv.writer(io.TextIOWrapper(sales_buffer, encoding='utf-8', newline=''))
        writer.writerow([
            'מספר חשבונית', 'לקוח', 'מיקום', 'סטטוס', 'סטטוס תשלום', 'אמצעי תשלום',
            'סכום ביניים', 'מע"מ', 'סה"כ', 'הערות', 'תאריך מכירה'
        ])
        sales = Sale.objects.select_related('customer', 'location').all()
        for sale in sales:
            writer.writerow([
                sale.invoice_number, sale.customer.name if sale.customer else '',
                sale.location.name if sale.location else '', sale.get_status_display(),
                sale.get_payment_status_display(),
                sale.get_payment_method_display() if sale.payment_method else '',
                sale.subtotal or 0, sale.tax_amount or 0, sale.total_amount or 0,
                sale.notes or '', sale.sale_date.strftime('%d/%m/%Y %H:%M') if sale.sale_date else ''
            ])
        zip_file.writestr('5_sales.csv', sales_buffer.getvalue())

        # 6. מיקומים
        locations_buffer = BytesIO()
        locations_buffer.write('\ufeff'.encode('utf-8'))
        writer = csv.writer(io.TextIOWrapper(locations_buffer, encoding='utf-8', newline=''))
        writer.writerow([
            'שם מיקום', 'סוג מיקום', 'כתובת', 'עיר', 'טלפון',
            'אחראי', 'כמות מוצרים', 'פעיל', 'תאריך יצירה'
        ])
        locations = Location.objects.annotate(product_count=Count('product_stocks'))
        for location in locations:
            writer.writerow([
                location.name, location.get_location_type_display(),
                location.address or '', location.city or '', location.phone or '',
                location.manager_name or '', location.product_count,
                'כן' if location.is_active else 'לא',
                location.created_at.strftime('%d/%m/%Y %H:%M') if location.created_at else ''
            ])
        zip_file.writestr('6_locations.csv', locations_buffer.getvalue())

        # 7. תנועות מלאי
        movements_buffer = BytesIO()
        movements_buffer.write('\ufeff'.encode('utf-8'))
        writer = csv.writer(io.TextIOWrapper(movements_buffer, encoding='utf-8', newline=''))
        writer.writerow([
            'מוצר', 'SKU', 'סוג תנועה', 'כמות', 'כמות קודמת', 'כמות חדשה',
            'סיבה', 'התייחסות', 'תאריך'
        ])
        movements = StockMovement.objects.select_related('product').all()
        for movement in movements:
            writer.writerow([
                movement.product.name, movement.product.sku or '',
                movement.get_movement_type_display(), movement.quantity,
                movement.previous_quantity, movement.new_quantity,
                movement.reason or '', movement.reference or '',
                movement.created_at.strftime('%d/%m/%Y %H:%M')
            ])
        zip_file.writestr('7_stock_movements.csv', movements_buffer.getvalue())

        # 8. התראות
        alerts_buffer = BytesIO()
        alerts_buffer.write('\ufeff'.encode('utf-8'))
        writer = csv.writer(io.TextIOWrapper(alerts_buffer, encoding='utf-8', newline=''))
        writer.writerow([
            'מוצר', 'סוג התראה', 'רמת חומרה', 'הודעה', 'נפתר', 'תאריך יצירה', 'תאריך פתרון'
        ])
        alerts = Alert.objects.select_related('product').all()
        for alert in alerts:
            writer.writerow([
                alert.product.name if alert.product else '',
                alert.get_alert_type_display(), alert.get_priority_display(),
                alert.message or '', 'כן' if alert.is_resolved else 'לא',
                alert.created_at.strftime('%d/%m/%Y %H:%M') if alert.created_at else '',
                alert.resolved_at.strftime('%d/%m/%Y %H:%M') if alert.resolved_at else ''
            ])
        zip_file.writestr('8_alerts.csv', alerts_buffer.getvalue())

    # החזרת קובץ ZIP
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="all_data_{timezone.now().strftime("%Y%m%d_%H%M%S")}.zip"'
    return response

def export_all_data_excel(request):
    """ייצוא כל הנתונים לקובץ Excel עם גליונות נפרדים"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # יצירת Workbook
    wb = Workbook()
    wb.remove(wb.active)  # הסרת הגליון הריק הראשוני

    # סגנון כותרות
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 1. גליון מוצרים
    ws_products = wb.create_sheet("מוצרים")
    ws_products.append(['שם', 'SKU', 'ברקוד', 'קטגוריה', 'כמות', 'מחיר מכירה', 'מחיר עלות', 'ערך מלאי', 'ספק'])
    products = Product.objects.select_related('category', 'supplier').all()
    for product in products:
        ws_products.append([
            product.name, product.sku or '', product.barcode or '',
            product.category.name if product.category else '',
            product.quantity, float(product.selling_price or 0), float(product.cost_price or 0),
            float(product.stock_value), product.supplier.name if product.supplier else ''
        ])

    # 2. גליון מכירות
    ws_sales = wb.create_sheet("מכירות")
    ws_sales.append(['מספר חשבונית', 'תאריך', 'לקוח', 'סכום כולל', 'סכום מס', 'מיקום', 'אמצעי תשלום', 'סטטוס'])
    sales = Sale.objects.select_related('customer', 'location').all()
    for sale in sales:
        ws_sales.append([
            sale.invoice_number, sale.created_at.strftime('%d/%m/%Y %H:%M'),
            sale.customer.name if sale.customer else '',
            float(sale.total_amount), float(sale.tax_amount or 0),
            sale.location.name if sale.location else '',
            sale.get_payment_method_display() if sale.payment_method else '',
            sale.get_status_display() if sale.status else ''
        ])

    # 3. גליון לקוחות
    ws_customers = wb.create_sheet("לקוחות")
    ws_customers.append(['שם', 'קוד לקוח', 'ח.פ/ע.מ', 'אימייל', 'טלפון', 'כתובת', 'הנחה %', 'מגבלת אשראי', 'סה"כ רכישות', 'פעיל'])
    customers = Customer.objects.all()
    for customer in customers:
        ws_customers.append([
            customer.name, customer.customer_code or '', customer.tax_id or '',
            customer.email or '', customer.phone or '', customer.address or '',
            float(customer.discount_percent or 0), float(customer.credit_limit or 0),
            float(customer.total_purchases or 0),
            'כן' if customer.is_active else 'לא'
        ])

    # 4. גליון קטגוריות
    ws_categories = wb.create_sheet("קטגוריות")
    ws_categories.append(['שם', 'תיאור', 'כמות מוצרים'])
    categories = Category.objects.annotate(product_count=Count('product')).all()
    for category in categories:
        ws_categories.append([
            category.name, category.description or '', category.product_count
        ])

    # 5. גליון ספקים
    ws_suppliers = wb.create_sheet("ספקים")
    ws_suppliers.append(['שם', 'איש קשר', 'אימייל', 'טלפון', 'כתובת', 'כמות מוצרים', 'פעיל'])
    suppliers = Supplier.objects.annotate(product_count=Count('product')).all()
    for supplier in suppliers:
        ws_suppliers.append([
            supplier.name, supplier.contact_person or '', supplier.email or '',
            supplier.phone or '', supplier.address or '', supplier.product_count,
            'כן' if supplier.is_active else 'לא'
        ])

    # 6. גליון מיקומים
    ws_locations = wb.create_sheet("מיקומים")
    ws_locations.append(['שם', 'סוג', 'כתובת', 'מנהל', 'כמות מוצרים', 'פעיל'])
    locations = Location.objects.annotate(product_count=Count('product_stocks')).all()
    for location in locations:
        ws_locations.append([
            location.name, location.get_location_type_display(), location.address or '',
            location.manager_name or '', location.product_count,
            'כן' if location.is_active else 'לא'
        ])

    # 7. גליון התראות
    ws_alerts = wb.create_sheet("התראות")
    ws_alerts.append(['מוצר', 'סוג התראה', 'עדיפות', 'הודעה', 'נפתר', 'תאריך יצירה'])
    alerts = Alert.objects.select_related('product').all()
    for alert in alerts:
        ws_alerts.append([
            alert.product.name if alert.product else '',
            alert.get_alert_type_display(), alert.get_priority_display(),
            alert.message or '', 'כן' if alert.is_resolved else 'לא',
            alert.created_at.strftime('%d/%m/%Y %H:%M') if alert.created_at else ''
        ])

    # עיצוב כל הגליונות
    for ws in wb.worksheets:
        # עיצוב כותרות
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # התאמת רוחב עמודות
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # הקפאת שורת כותרות
        ws.freeze_panes = 'A2'

    # שמירה ל-BytesIO
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # החזרת התגובה
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="inventory_data_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response

# Views למכירות
def customer_list(request):
    """רשימת לקוחות"""
    customers = Customer.objects.all()
    return render(request, 'inventory/customer_list.html', {'customers': customers})

def add_customer(request):
    """הוספת לקוח חדש"""
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            try:
                customer = form.save(commit=False)
                # רק אם המשתמש מחובר, נשמור אותו
                if request.user.is_authenticated:
                    customer.created_by = request.user
                customer.save()
                messages.success(request, 'הלקוח נוסף בהצלחה!')
                return redirect('customer_list')
            except Exception as e:
                messages.error(request, f'שגיאה בשמירת הלקוח: {str(e)}')
        else:
            messages.error(request, 'יש שגיאות בטופס. אנא בדוק את הנתונים.')
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = CustomerForm()
    return render(request, 'inventory/add_customer.html', {'form': form})

def sale_list(request):
    """רשימת מכירות"""
    sales = Sale.objects.select_related('customer').all()
    return render(request, 'inventory/sale_list.html', {'sales': sales})

def add_sale(request):
    """הוספת מכירה חדשה עם אפשרות להוסיף לקוח מהר"""
    if request.method == 'POST':
        # בדיקה אם מוסיפים לקוח חדש במהירות
        if 'add_quick_customer' in request.POST:
            customer_name = request.POST.get('quick_customer_name', '').strip()
            customer_phone = request.POST.get('quick_customer_phone', '').strip()
            customer_email = request.POST.get('quick_customer_email', '').strip()

            if customer_name:
                # יצירת לקוח חדש
                customer = Customer.objects.create(
                    name=customer_name,
                    phone=customer_phone if customer_phone else None,
                    email=customer_email if customer_email else None,
                    customer_type='individual',
                    created_by=request.user if request.user.is_authenticated else None
                )
                messages.success(request, f'הלקוח "{customer_name}" נוסף בהצלחה!')

                # החזרת הטופס עם הלקוח החדש נבחר
                form = SaleForm(initial={'customer': customer.id})
                formset = SaleItemFormSet()
                return render(request, 'inventory/add_sale.html', {
                    'form': form,
                    'formset': formset,
                    'new_customer_id': customer.id
                })
            else:
                messages.error(request, 'נא להזין שם לקוח!')
                form = SaleForm()
                formset = SaleItemFormSet()
                return render(request, 'inventory/add_sale.html', {'form': form, 'formset': formset})

        # טיפול במכירה רגילה
        form = SaleForm(request.POST)
        formset = SaleItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            # בדיקת מלאי לפני שמירה
            from .models import ProductLocationStock
            from django.utils import timezone

            # קבל את המיקום מהטופס
            location = form.cleaned_data.get('location')

            # בדוק כל פריט שאין חריגה במלאי
            stock_errors = []
            for item_form in formset:
                if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE', False):
                    product = item_form.cleaned_data.get('product')
                    quantity = item_form.cleaned_data.get('quantity', 0)

                    if product and quantity > 0:
                        # בדוק מלאי במיקום אם צוין
                        if location:
                            location_stock = ProductLocationStock.objects.filter(
                                product=product,
                                location=location
                            ).first()

                            available_quantity = location_stock.quantity if location_stock else 0

                            if quantity > available_quantity:
                                stock_errors.append(
                                    f'❌ {product.name}: מבוקש {quantity} יחידות, זמין במלאי: {available_quantity} יחידות ב{location.name}'
                                )
                        else:
                            # בדוק מלאי כללי אם אין מיקום
                            if quantity > product.quantity:
                                stock_errors.append(
                                    f'❌ {product.name}: מבוקש {quantity} יחידות, זמין במלאי: {product.quantity} יחידות'
                                )

            # אם יש שגיאות מלאי, הצג אותן ולא תשמור
            if stock_errors:
                for error in stock_errors:
                    messages.error(request, error)
                messages.warning(request, '⚠️ המכירה לא נשמרה - אין מספיק מלאי למוצרים המבוקשים!')

                # החזר את הטופס עם הנתונים שהמשתמש הזין
                import json
                products_data = {}
                for product in Product.objects.all():
                    products_data[str(product.id)] = {
                        'name': product.name,
                        'selling_price': float(product.selling_price or 0),
                        'quantity': product.quantity
                    }

                customers_data = {}
                for customer in Customer.objects.all():
                    customers_data[str(customer.id)] = {
                        'name': customer.name,
                        'discount': float(customer.discount_percent or 0)
                    }

                context = {
                    'form': form,
                    'formset': formset,
                    'customers': Customer.objects.all().order_by('name'),
                    'products': Product.objects.filter(quantity__gt=0).order_by('name'),
                    'locations': Location.objects.filter(is_active=True).order_by('location_type', 'name'),
                    'products_json': json.dumps(products_data),
                    'customers_json': json.dumps(customers_data),
                }
                return render(request, 'inventory/add_sale.html', context)

            # אם אין שגיאות מלאי, המשך עם השמירה
            with transaction.atomic():
                sale = form.save(commit=False)
                if request.user.is_authenticated:
                    sale.created_by = request.user
                # מספר החשבונית ייווצר אוטומטית ב-save
                sale.save()

                # שמירת פריטי החשבונית
                formset.instance = sale
                sale_items = formset.save()

                # עדכון סכום כולל
                # total_price כבר כולל מע"מ, צריך להפריד אותו
                total_with_vat = sum(item.total_price for item in sale.items.all())
                # חישוב subtotal (לפני מע"מ) מתוך total (כולל מע"מ)
                sale.subtotal = total_with_vat / (1 + sale.tax_rate / 100)
                sale.save()

                # עדכון מלאי - במיקום הספציפי ובסה"כ כללי
                for item in sale_items:
                    product = item.product

                    # אם יש מיקום למכירה, עדכן מלאי במיקום
                    if sale.location:
                        # מצא או צור רשומת מלאי למיקום זה
                        location_stock, created = ProductLocationStock.objects.get_or_create(
                            product=product,
                            location=sale.location,
                            defaults={'quantity': 0}
                        )

                        # עדכן מלאי במיקום
                        location_stock.quantity -= item.quantity
                        location_stock.last_sold = timezone.now()
                        location_stock.save()

                    # עדכן גם את המלאי הכללי במוצר
                    product.quantity -= item.quantity
                    product.save()

                messages.success(request, f'✅ המכירה נוצרה בהצלחה! מספר חשבונית: {sale.invoice_number}')
                return redirect('sale_detail', pk=sale.pk)
        else:
            messages.error(request, 'אנא תקן את השגיאות בטופס.')
    else:
        # קבע חנות ראשית כברירת מחדל
        default_store = Location.objects.filter(
            location_type='store',
            is_main_store=True
        ).first()

        if not default_store:
            # אם אין חנות ראשית, קח את החנות הראשונה
            default_store = Location.objects.filter(location_type='store').first()

        form = SaleForm(initial={'location': default_store.id if default_store else None})
        formset = SaleItemFormSet()

    # הכן נתוני מוצרים ולקוחות עבור JavaScript
    import json
    products_data = {}
    for product in Product.objects.filter(quantity__gt=0):
        products_data[str(product.id)] = {
            'name': product.name,
            'selling_price': float(product.selling_price or 0),
            'quantity': product.quantity
        }

    customers_data = {}
    for customer in Customer.objects.all():
        customers_data[str(customer.id)] = {
            'name': customer.name,
            'discount': float(customer.discount_percent or 0)
        }

    context = {
        'form': form,
        'formset': formset,
        'customers': Customer.objects.all().order_by('name'),
        'products': Product.objects.filter(quantity__gt=0).order_by('name'),
        'locations': Location.objects.filter(is_active=True).order_by('location_type', 'name'),
        'products_json': json.dumps(products_data),
        'customers_json': json.dumps(customers_data),
    }
    return render(request, 'inventory/add_sale.html', context)

def sale_detail(request, pk):
    """פרטי מכירה"""
    sale = get_object_or_404(Sale, pk=pk)
    return render(request, 'inventory/sale_detail.html', {'sale': sale})

def edit_sale(request, pk):
    """עריכת מכירה"""
    sale = get_object_or_404(Sale, pk=pk)

    if request.method == 'POST':
        form = SaleForm(request.POST, instance=sale)
        formset = SaleItemFormSet(request.POST, instance=sale)

        if form.is_valid() and formset.is_valid():
            # בדיקת מלאי לפני עדכון
            from .models import ProductLocationStock
            from django.utils import timezone

            location = form.cleaned_data.get('location')

            # החזר מלאי ישן זמנית למטרת הבדיקה
            old_items = {}
            for item in sale.items.all():
                old_items[item.product.id] = item.quantity

            # בדוק מלאי לפריטים החדשים
            stock_errors = []
            for item_form in formset:
                if item_form.cleaned_data and not item_form.cleaned_data.get('DELETE', False):
                    product = item_form.cleaned_data.get('product')
                    new_quantity = item_form.cleaned_data.get('quantity', 0)

                    if product and new_quantity > 0:
                        # כמות ישנה (אם הייתה)
                        old_quantity = old_items.get(product.id, 0)
                        # הפרש בין חדש לישן
                        quantity_diff = new_quantity - old_quantity

                        # אם מוסיפים מלאי (quantity_diff > 0), בדוק שיש מספיק
                        if quantity_diff > 0:
                            if location:
                                location_stock = ProductLocationStock.objects.filter(
                                    product=product,
                                    location=location
                                ).first()

                                available_quantity = location_stock.quantity if location_stock else 0

                                if quantity_diff > available_quantity:
                                    stock_errors.append(
                                        f'❌ {product.name}: נדרש להוסיף {quantity_diff} יחידות, זמין במלאי: {available_quantity} יחידות ב{location.name}'
                                    )
                            else:
                                if quantity_diff > product.quantity:
                                    stock_errors.append(
                                        f'❌ {product.name}: נדרש להוסיף {quantity_diff} יחידות, זמין במלאי: {product.quantity} יחידות'
                                    )

            # אם יש שגיאות מלאי, הצג אותן
            if stock_errors:
                for error in stock_errors:
                    messages.error(request, error)
                messages.warning(request, '⚠️ המכירה לא עודכנה - אין מספיק מלאי למוצרים המבוקשים!')

                context = {
                    'form': form,
                    'formset': formset,
                    'sale': sale,
                    'is_edit': True,
                }
                return render(request, 'inventory/add_sale.html', context)

            # אם אין שגיאות, המשך עם העדכון
            with transaction.atomic():
                # החזר מלאי ישן לפני עדכון
                for item in sale.items.all():
                    product = item.product
                    product.quantity += item.quantity
                    if sale.location:
                        location_stock = ProductLocationStock.objects.filter(
                            product=product,
                            location=sale.location
                        ).first()
                        if location_stock:
                            location_stock.quantity += item.quantity
                            location_stock.save()
                    product.save()

                # עדכן מכירה
                sale = form.save()
                formset.instance = sale
                sale_items = formset.save()

                # עדכן סכום
                # total_price כבר כולל מע"מ, צריך להפריד אותו
                total_with_vat = sum(item.total_price for item in sale.items.all())
                sale.subtotal = total_with_vat / (1 + sale.tax_rate / 100)
                sale.save()

                # הורד מלאי חדש
                for item in sale_items:
                    product = item.product
                    if sale.location:
                        location_stock, created = ProductLocationStock.objects.get_or_create(
                            product=product,
                            location=sale.location,
                            defaults={'quantity': 0}
                        )
                        location_stock.quantity -= item.quantity
                        location_stock.last_sold = timezone.now()
                        location_stock.save()
                    product.quantity -= item.quantity
                    product.save()

                messages.success(request, '✅ המכירה עודכנה בהצלחה!')
                return redirect('sale_detail', pk=sale.pk)
    else:
        form = SaleForm(instance=sale)
        formset = SaleItemFormSet(instance=sale)

    context = {
        'form': form,
        'formset': formset,
        'sale': sale,
        'is_edit': True,
    }
    return render(request, 'inventory/add_sale.html', context)

def delete_sale(request, pk):
    """מחיקת מכירה"""
    sale = get_object_or_404(Sale, pk=pk)

    if request.method == 'POST':
        with transaction.atomic():
            # החזר מלאי
            for item in sale.items.all():
                product = item.product
                product.quantity += item.quantity
                if sale.location:
                    location_stock = ProductLocationStock.objects.filter(
                        product=product,
                        location=sale.location
                    ).first()
                    if location_stock:
                        location_stock.quantity += item.quantity
                        location_stock.save()
                product.save()

            invoice_num = sale.invoice_number
            sale.delete()
            messages.success(request, f'חשבונית {invoice_num} נמחקה והמלאי הוחזר')
            return redirect('sale_list')

    return render(request, 'inventory/delete_sale.html', {'sale': sale})

def search_product_by_barcode(request):
    """API endpoint לחיפוש מוצר לפי ברקוד"""
    barcode = request.GET.get('barcode', '').strip()

    if not barcode:
        return JsonResponse({'error': 'ברקוד לא סופק'}, status=400)

    try:
        product = Product.objects.get(barcode=barcode, status='active')

        data = {
            'success': True,
            'product': {
                'id': product.id,
                'name': product.name,
                'sku': product.sku or '',
                'barcode': product.barcode,
                'price': float(product.selling_price or product.price or 0),
                'quantity': product.quantity,
                'unit': product.get_unit_display(),
                'category': product.category.name if product.category else '',
            }
        }
        return JsonResponse(data)

    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'מוצר לא נמצא',
            'message': f'לא נמצא מוצר עם ברקוד {barcode}'
        }, status=404)

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': 'שגיאה בחיפוש',
            'message': str(e)
        }, status=500)

# Views להעברות מלאי
def transfer_list(request):
    """רשימת העברות מלאי"""
    transfers = InventoryTransfer.objects.select_related('from_location', 'to_location').all()
    return render(request, 'inventory/transfer_list.html', {'transfers': transfers})

def add_transfer(request):
    """הוספת העברת מלאי חדשה"""
    if request.method == 'POST':
        form = InventoryTransferForm(request.POST)
        formset = TransferItemFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                transfer = form.save(commit=False)
                # יצירת מספר העברה אוטומטי
                if not transfer.transfer_number:
                    transfer.transfer_number = f"TRF-{InventoryTransfer.objects.count() + 1:06d}"
                transfer.save()

                # שמירת פריטי ההעברה
                formset.instance = transfer
                formset.save()

                # עדכון מלאי אוטומטי אם ההעברה הושלמה
                if transfer.status == 'completed':
                    update_inventory_from_transfer(transfer)

                messages.success(request, 'ההעברה נוספה בהצלחה!')
                return redirect('transfer_detail', pk=transfer.pk)
    else:
        form = InventoryTransferForm()
        formset = TransferItemFormSet()

    return render(request, 'inventory/add_transfer.html', {'form': form, 'formset': formset})

def transfer_detail(request, pk):
    """פרטי העברה"""
    transfer = get_object_or_404(InventoryTransfer, pk=pk)
    return render(request, 'inventory/transfer_detail.html', {'transfer': transfer})

# מיקומים
def location_list(request):
    """רשימת מיקומים"""
    locations = Location.objects.all()
    return render(request, 'inventory/location_list.html', {'locations': locations})

def add_location(request):
    """הוספת מיקום חדש"""
    if request.method == 'POST':
        form = LocationForm(request.POST)
        if form.is_valid():
            try:
                location = form.save(commit=False)
                location.created_by = request.user
                location.save()
                messages.success(request, 'המיקום נוסף בהצלחה!')
                return redirect('location_list')
            except Exception as e:
                messages.error(request, f'שגיאה בשמירת המיקום: {str(e)}')
        else:
            messages.error(request, 'יש שגיאות בטופס. אנא בדוק את הנתונים.')
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = LocationForm()
    return render(request, 'inventory/add_location.html', {'form': form})

def location_detail(request, pk):
    """פרטי מיקום"""
    location = get_object_or_404(Location, pk=pk)
    products = Product.objects.filter(location=location)
    alerts = LocationAlert.objects.filter(location=location, is_resolved=False)
    return render(request, 'inventory/location_detail.html', {
        'location': location,
        'products': products,
        'alerts': alerts
    })

def edit_location(request, pk):
    """עריכת מיקום"""
    location = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        form = LocationForm(request.POST, instance=location)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'המיקום עודכן בהצלחה!')
                return redirect('location_detail', pk=pk)
            except Exception as e:
                messages.error(request, f'שגיאה בעדכון המיקום: {str(e)}')
        else:
            messages.error(request, 'יש שגיאות בטופס. אנא בדוק את הנתונים.')
    else:
        form = LocationForm(instance=location)
    return render(request, 'inventory/edit_location.html', {'form': form, 'location': location})

def delete_location(request, pk):
    """מחיקת מיקום"""
    location = get_object_or_404(Location, pk=pk)
    if request.method == 'POST':
        location_name = location.name
        location.delete()
        messages.success(request, f'המיקום "{location_name}" נמחק בהצלחה!')
        return redirect('location_list')
    return render(request, 'inventory/delete_location.html', {'location': location})

def complete_transfer(request, pk):
    """השלמת העברה ועדכון מלאי"""
    transfer = get_object_or_404(InventoryTransfer, pk=pk)

    if transfer.status != 'completed':
        with transaction.atomic():
            transfer.status = 'completed'
            transfer.save()

            # עדכון מלאי
            update_inventory_from_transfer(transfer)

            messages.success(request, 'ההעברה הושלמה והמלאי עודכן!')

    return redirect('transfer_detail', pk=transfer.pk)

def update_inventory_from_transfer(transfer):
    """עדכון מלאי אוטומטי מהעברה"""
    for item in transfer.items.all():
        # הפחתה מהמיקום המקור
        if item.product.location == transfer.from_location:
            item.product.quantity = max(0, item.product.quantity - item.quantity)
        else:
            # אם המוצר לא במיקום המקור, נחפש אותו
            try:
                product_in_source = Product.objects.get(
                    name=item.product.name,
                    location=transfer.from_location
                )
                product_in_source.quantity = max(0, product_in_source.quantity - item.quantity)
                product_in_source.save()
            except Product.DoesNotExist:
                pass

        # הוספה למיקום היעד
        if item.product.location == transfer.to_location:
            item.product.quantity += item.quantity
        else:
            # אם המוצר לא במיקום היעד, נחפש אותו או ניצור חדש
            try:
                product_in_dest = Product.objects.get(
                    name=item.product.name,
                    location=transfer.to_location
                )
                product_in_dest.quantity += item.quantity
                product_in_dest.save()
            except Product.DoesNotExist:
                # יצירת עותק של המוצר במיקום היעד
                new_product = Product.objects.create(
                    name=item.product.name,
                    description=item.product.description,
                    category=item.product.category,
                    supplier=item.product.supplier,
                    location=transfer.to_location,
                    quantity=item.quantity,
                    min_quantity=item.product.min_quantity,
                    max_quantity=item.product.max_quantity,
                    unit=item.product.unit,
                    cost_price=item.product.cost_price,
                    selling_price=item.product.selling_price,
                    barcode=item.product.barcode,
                    status=item.product.status,
                    weight=item.product.weight,
                    dimensions=item.product.dimensions,
                    notes=item.product.notes,
                )

        item.product.save()

        # יצירת תנועת מלאי
        StockMovement.objects.create(
            product=item.product,
            movement_type='transfer_out' if transfer.transfer_type.startswith('warehouse_to') else 'transfer_in',
            quantity=-item.quantity if transfer.transfer_type.startswith('warehouse_to') else item.quantity,
            reason=f'העברה #{transfer.transfer_number}',
        )

# Views ל-CRM
def crm_dashboard(request):
    """דשבורד CRM"""
    # סטטיסטיקות לקוחות
    total_customers = Customer.objects.count()
    active_customers = Customer.objects.filter(is_active=True).count()
    vip_customers = Customer.objects.filter(is_vip=True).count()

    # לקוחות לפי סוג
    customers_by_type = Customer.objects.values('customer_type').annotate(
        count=Count('id')
    ).order_by('-count')

    # לקוחות לפי עדיפות
    customers_by_priority = Customer.objects.values('priority').annotate(
        count=Count('id')
    ).order_by('-count')

    # לקוחות עם רכישות גבוהות
    top_customers = Customer.objects.filter(
        total_purchases__gt=0
    ).order_by('-total_purchases')[:10]

    # התראות לקוחות
    customer_alerts = CustomerAlert.objects.filter(
        is_resolved=False
    ).order_by('-created_at')[:10]

    # הערות אחרונות
    recent_notes = CustomerNote.objects.select_related(
        'customer', 'created_by'
    ).order_by('-created_at')[:10]

    # מכירות אחרונות
    recent_sales = Sale.objects.select_related('customer').order_by('-created_at')[:10]

    context = {
        'total_customers': total_customers,
        'active_customers': active_customers,
        'vip_customers': vip_customers,
        'customers_by_type': customers_by_type,
        'customers_by_priority': customers_by_priority,
        'top_customers': top_customers,
        'customer_alerts': customer_alerts,
        'recent_notes': recent_notes,
        'recent_sales': recent_sales,
    }
    return render(request, 'inventory/crm_dashboard.html', context)

def customer_detail(request, pk):
    """פרטי לקוח"""
    customer = get_object_or_404(Customer, pk=pk)

    # היסטוריית מכירות
    sale_history = CustomerSaleHistory.objects.filter(
        customer=customer
    ).order_by('-sale_date')[:20]

    # הערות לקוח
    notes = CustomerNote.objects.filter(
        customer=customer
    ).order_by('-created_at')[:10]

    # התראות לקוח
    alerts = CustomerAlert.objects.filter(
        customer=customer
    ).order_by('-created_at')[:10]

    # סטטיסטיקות
    total_sales = sale_history.count()
    total_amount = sale_history.aggregate(
        total=Sum('total_amount')
    )['total'] or 0

    context = {
        'customer': customer,
        'sale_history': sale_history,
        'notes': notes,
        'alerts': alerts,
        'total_sales': total_sales,
        'total_amount': total_amount,
    }
    return render(request, 'inventory/customer_detail.html', context)

def edit_customer(request, pk):
    """עריכת לקוח"""
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, 'הלקוח עודכן בהצלחה!')
            return redirect('customer_detail', pk=customer.pk)
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'inventory/edit_customer.html', {'form': form, 'customer': customer})

def delete_customer(request, pk):
    """מחיקת לקוח"""
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        customer_name = customer.name
        customer.delete()
        messages.success(request, f'הלקוח "{customer_name}" נמחק בהצלחה!')
        return redirect('customer_list')
    return render(request, 'inventory/delete_customer.html', {'customer': customer})

def add_customer_note(request, customer_pk):
    """הוספת הערה ללקוח"""
    customer = get_object_or_404(Customer, pk=customer_pk)

    if request.method == 'POST':
        form = CustomerNoteForm(request.POST)
        if form.is_valid():
            note = form.save(commit=False)
            note.customer = customer
            note.created_by = request.user
            note.save()
            messages.success(request, 'ההערה נוספה בהצלחה!')
            return redirect('customer_detail', pk=customer.pk)
    else:
        form = CustomerNoteForm()

    return render(request, 'inventory/add_customer_note.html', {
        'form': form, 'customer': customer
    })

def add_customer_alert(request, customer_pk):
    """הוספת התראה ללקוח"""
    customer = get_object_or_404(Customer, pk=customer_pk)

    if request.method == 'POST':
        form = CustomerAlertForm(request.POST)
        if form.is_valid():
            alert = form.save(commit=False)
            alert.customer = customer
            alert.save()
            messages.success(request, 'ההתראה נוספה בהצלחה!')
            return redirect('customer_detail', pk=customer.pk)
    else:
        form = CustomerAlertForm()

    return render(request, 'inventory/add_customer_alert.html', {
        'form': form, 'customer': customer
    })

def generate_invoice(request, sale_pk):
    """הפקת חשבונית מס רשמית לפי חוקי המס בישראל"""
    sale = get_object_or_404(Sale, pk=sale_pk)

    # קבלת תבנית ברירת מחדל
    template = InvoiceTemplate.objects.filter(is_default=True).first()
    if not template:
        template = InvoiceTemplate.objects.first()

    context = {
        'sale': sale,
        'template': template,
    }
    return render(request, 'inventory/tax_invoice.html', context)

def invoice_templates(request):
    """ניהול תבניות חשבוניות"""
    templates = InvoiceTemplate.objects.all()
    return render(request, 'inventory/invoice_templates.html', {'templates': templates})

def add_invoice_template(request):
    """הוספת תבנית חשבונית"""
    if request.method == 'POST':
        form = InvoiceTemplateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'התבנית נוספה בהצלחה!')
            return redirect('invoice_templates')
    else:
        form = InvoiceTemplateForm()

    return render(request, 'inventory/add_invoice_template.html', {'form': form})

def customer_reports(request):
    """דוחות לקוחות"""
    # סטטיסטיקות כלליות
    total_customers = Customer.objects.count()
    active_customers = Customer.objects.filter(is_active=True).count()
    vip_customers = Customer.objects.filter(is_vip=True).count()

    # לקוחות לפי סוג
    customers_by_type = Customer.objects.values('customer_type').annotate(
        count=Count('id')
    ).order_by('-count')

    # לקוחות לפי עיר
    customers_by_city = Customer.objects.values('city').annotate(
        count=Count('id')
    ).order_by('-count')[:20]

    # לקוחות עם הרכישות הגבוהות ביותר (על בסיס מכירות)
    top_customers = Customer.objects.annotate(
        purchase_count=Count('sale'),
        total_spent=Sum('sale__total_amount')
    ).filter(
        purchase_count__gt=0
    ).order_by('-total_spent')[:20]

    # לקוחות לא פעילים
    inactive_customers = Customer.objects.filter(
        is_active=False
    ).order_by('-created_at')

    # לקוחות ללא רכישות
    customers_without_purchases = Customer.objects.annotate(
        purchase_count=Count('sale')
    ).filter(
        purchase_count=0
    ).order_by('-created_at')

    # המרת נתונים ל-JSON עבור הגרפים
    import json
    customers_by_type_json = json.dumps(list(customers_by_type))
    customers_by_city_json = json.dumps(list(customers_by_city))

    context = {
        'total_customers': total_customers,
        'active_customers': active_customers,
        'vip_customers': vip_customers,
        'customers_by_type': customers_by_type_json,
        'customers_by_city': customers_by_city_json,
        'top_customers': top_customers,
        'inactive_customers': inactive_customers,
        'customers_without_purchases': customers_without_purchases,
    }
    return render(request, 'inventory/customer_reports.html', context)

# גיבוי ושחזור נתונים
import os
import subprocess
from django.conf import settings
from django.views.decorators.http import require_http_methods

@require_http_methods(["POST"])
def backup_data(request):
    """יצירת גיבוי של המערכת"""
    try:
        # הרצת פקודת Django management command
        result = subprocess.run(
            ['python', 'manage.py', 'backup_database'],
            capture_output=True,
            text=True,
            cwd=settings.BASE_DIR
        )

        if result.returncode == 0:
            # חיפוש הגיבוי האחרון שנוצר
            backup_dir = os.path.join(settings.BASE_DIR, 'backups')
            if os.path.exists(backup_dir):
                backups = sorted(
                    [f for f in os.listdir(backup_dir) if f.endswith('.zip')],
                    key=lambda x: os.path.getmtime(os.path.join(backup_dir, x)),
                    reverse=True
                )
                if backups:
                    return JsonResponse({
                        'success': True,
                        'backup_file': backups[0],
                        'message': 'הגיבוי נוצר בהצלחה!'
                    })

            return JsonResponse({
                'success': True,
                'message': 'הגיבוי נוצר בהצלחה!'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.stderr or 'שגיאה לא ידועה ביצירת הגיבוי'
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

def list_backups(request):
    """רשימת גיבויים זמינים"""
    try:
        backup_dir = os.path.join(settings.BASE_DIR, 'backups')
        if not os.path.exists(backup_dir):
            return JsonResponse({'backups': []})

        backups = []
        for filename in os.listdir(backup_dir):
            if filename.endswith('.zip'):
                file_path = os.path.join(backup_dir, filename)
                file_stats = os.stat(file_path)
                backups.append({
                    'name': filename,
                    'size': f"{file_stats.st_size / 1024:.1f} KB",
                    'date': datetime.fromtimestamp(file_stats.st_mtime).strftime('%d/%m/%Y %H:%M')
                })

        # מיון לפי תאריך (החדש ביותר ראשון)
        backups.sort(key=lambda x: x['date'], reverse=True)

        return JsonResponse({'backups': backups})
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@require_http_methods(["POST"])
def restore_data(request):
    """שחזור נתונים מגיבוי"""
    try:
        data = json.loads(request.body)
        backup_file = data.get('backup_file')

        if not backup_file:
            return JsonResponse({
                'success': False,
                'error': 'לא צוין קובץ גיבוי'
            })

        # בדיקה שהקובץ קיים
        backup_path = os.path.join(settings.BASE_DIR, 'backups', backup_file)
        if not os.path.exists(backup_path):
            return JsonResponse({
                'success': False,
                'error': 'קובץ הגיבוי לא נמצא'
            })

        # הרצת פקודת שחזור
        result = subprocess.run(
            ['python', 'manage.py', 'restore_database', backup_file],
            capture_output=True,
            text=True,
            cwd=settings.BASE_DIR
        )

        if result.returncode == 0:
            return JsonResponse({
                'success': True,
                'message': 'הגיבוי שוחזר בהצלחה!'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.stderr or 'שגיאה לא ידועה בשחזור הגיבוי'
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

# הגדרות מערכת
def system_settings(request):
    """עריכת הגדרות מערכת"""
    settings_obj = SystemSettings.load()

    if request.method == 'POST':
        form = SystemSettingsForm(request.POST, instance=settings_obj)
        if form.is_valid():
            settings_instance = form.save(commit=False)
            if request.user.is_authenticated:
                settings_instance.updated_by = request.user
            settings_instance.save()
            messages.success(request, '✅ ההגדרות נשמרו בהצלחה!')
            return redirect('system_settings')
        else:
            messages.error(request, '❌ שגיאה בשמירת ההגדרות. בדוק את השדות.')
    else:
        form = SystemSettingsForm(instance=settings_obj)

    context = {
        'form': form,
        'settings': settings_obj,
    }
    return render(request, 'inventory/system_settings.html', context)

def test_email_settings(request):
    """שליחת מייל בדיקה"""
    if request.method == 'POST':
        try:
            settings_obj = SystemSettings.load()

            if not settings_obj.email_enabled:
                return JsonResponse({
                    'success': False,
                    'error': 'שליחת מיילים לא מופעלת במערכת. הפעל ב"הגדרות"'
                })

            # עדכון הגדרות Django מהמסד נתונים
            from django.core.mail import EmailMessage
            from django.conf import settings

            settings.EMAIL_HOST = settings_obj.email_host
            settings.EMAIL_PORT = settings_obj.email_port
            settings.EMAIL_USE_TLS = settings_obj.email_use_tls
            settings.EMAIL_USE_SSL = settings_obj.email_use_ssl
            settings.EMAIL_HOST_USER = settings_obj.email_host_user
            settings.EMAIL_HOST_PASSWORD = settings_obj.email_host_password
            settings.DEFAULT_FROM_EMAIL = settings_obj.default_from_email or settings_obj.email_host_user

            # יצירת מייל בדיקה
            email = EmailMessage(
                subject='🧪 בדיקת הגדרות Email - מערכת ניהול מלאי',
                body=f"""
                <html dir="rtl">
                <body style="font-family: Arial; padding: 20px;">
                    <h2 style="color: #28a745;">✅ מייל בדיקה</h2>
                    <p>אם אתה מקבל מייל זה, ההגדרות שלך נכונות!</p>
                    <hr>
                    <p><strong>פרטי השרת:</strong></p>
                    <ul>
                        <li>שרת: {settings_obj.email_host}</li>
                        <li>פורט: {settings_obj.email_port}</li>
                        <li>TLS: {'כן' if settings_obj.email_use_tls else 'לא'}</li>
                        <li>שולח: {settings_obj.email_host_user}</li>
                    </ul>
                    <p style="color: #6c757d; font-size: 12px;">נשלח מ: מערכת ניהול מלאי</p>
                </body>
                </html>
                """,
                from_email=settings_obj.default_from_email or settings_obj.email_host_user,
                to=[settings_obj.daily_report_email] if settings_obj.daily_report_email else [settings_obj.email_host_user],
            )
            email.content_subtype = 'html'

            # שליחת המייל
            import ssl
            # כדי לטפל בבעיות SSL עם תעודות self-signed
            try:
                email.send()
            except Exception as ssl_error:
                # נסה שוב עם אימות SSL מושבת
                import smtplib
                if 'CERTIFICATE_VERIFY_FAILED' in str(ssl_error):
                    # צור קונטקסט SSL שמתעלם מאימות תעודות
                    context = ssl.create_default_context()
                    context.check_hostname = False
                    context.verify_mode = ssl.CERT_NONE
                    
                    # שלח ידנית עם ההגדרות המותאמות
                    connection = smtplib.SMTP(settings_obj.email_host, settings_obj.email_port)
                    if settings_obj.email_use_tls:
                        connection.starttls(context=context)
                    connection.login(settings_obj.email_host_user, settings_obj.email_host_password)
                    connection.send_message(email.message())
                    connection.quit()
                else:
                    raise ssl_error

            return JsonResponse({
                'success': True,
                'message': f'✅ מייל בדיקה נשלח ל-{settings_obj.daily_report_email or settings_obj.email_host_user}!'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'❌ שגיאה בשליחת המייל: {str(e)}'
            })

    return JsonResponse({'success': False, 'error': 'Method not allowed'})

def send_instant_report(request):
    """שליחת דוח מיידי ללא תזמון"""
    if request.method == 'POST':
        try:
            settings_obj = SystemSettings.load()

            # בדיקות ראשוניות
            if not settings_obj.email_enabled:
                return JsonResponse({
                    'success': False,
                    'error': 'שליחת מיילים לא מופעלת. הפעל ב"הגדרות"'
                })

            if not settings_obj.daily_report_email:
                return JsonResponse({
                    'success': False,
                    'error': 'לא הוגדר מייל לקבלת דוחות. הגדר ב"הגדרות"'
                })

            # עדכון הגדרות Django מהמסד נתונים
            from django.conf import settings
            settings.EMAIL_HOST = settings_obj.email_host
            settings.EMAIL_PORT = settings_obj.email_port
            settings.EMAIL_USE_TLS = settings_obj.email_use_tls
            settings.EMAIL_USE_SSL = settings_obj.email_use_ssl
            settings.EMAIL_HOST_USER = settings_obj.email_host_user
            settings.EMAIL_HOST_PASSWORD = settings_obj.email_host_password
            settings.DEFAULT_FROM_EMAIL = settings_obj.default_from_email or settings_obj.email_host_user

            # איסוף נתונים
            today = timezone.now().date()

            # סטטיסטיקות כלליות
            total_products = Product.objects.count()
            low_stock = Product.objects.filter(quantity__lte=F('min_quantity')).count()
            out_of_stock = Product.objects.filter(quantity=0).count()
            total_stock_value = Product.objects.aggregate(
                total=Sum(F('quantity') * F('cost_price'))
            )['total'] or 0

            # מכירות היום
            today_sales = Sale.objects.filter(created_at__date=today)
            daily_sales_count = today_sales.count()
            daily_sales_amount = today_sales.aggregate(total=Sum('total_amount'))['total'] or 0

            # התראות
            active_alerts = Alert.objects.filter(is_resolved=False).count()
            critical_alerts = Alert.objects.filter(is_resolved=False, priority='critical').count()

            # תנועות מלאי
            today_movements = StockMovement.objects.filter(created_at__date=today).count()

            # מוצרים במלאי נמוך
            low_stock_products = Product.objects.filter(
                quantity__lte=F('min_quantity')
            ).order_by('quantity')[:10]

            # יצירת תוכן HTML
            html_content = f"""
            <!DOCTYPE html>
            <html dir="rtl" lang="he">
            <head>
                <meta charset="UTF-8">
                <style>
                    body {{ font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 20px; }}
                    .container {{ max-width: 800px; margin: 0 auto; background: white; padding: 30px; border-radius: 10px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
                    h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
                    h2 {{ color: #34495e; margin-top: 30px; border-right: 4px solid #3498db; padding-right: 10px; }}
                    .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 15px; margin: 20px 0; }}
                    .stat-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; }}
                    .stat-card.warning {{ background: linear-gradient(135deg, #f093fb 0%, #f5576c 100%); }}
                    .stat-card.success {{ background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%); }}
                    .stat-value {{ font-size: 28px; font-weight: bold; margin: 10px 0; }}
                    .stat-label {{ font-size: 13px; opacity: 0.9; }}
                    table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                    th {{ background-color: #3498db; color: white; padding: 10px; text-align: right; }}
                    td {{ padding: 8px; border-bottom: 1px solid #ddd; text-align: right; }}
                    .alert {{ padding: 12px; margin: 10px 0; border-radius: 5px; }}
                    .alert-danger {{ background-color: #f8d7da; border-right: 4px solid #dc3545; color: #721c24; }}
                    .footer {{ margin-top: 30px; padding-top: 20px; border-top: 2px solid #eee; text-align: center; color: #7f8c8d; font-size: 12px; }}
                    .instant {{ background-color: #fff3cd; border: 2px solid #ffc107; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                </style>
            </head>
            <body>
                <div class="container">
                    <div class="instant">
                        <strong>⚡ דוח מיידי</strong> - נשלח לבקשתך ב-{timezone.now().strftime('%H:%M')}
                    </div>

                    <h1>📊 דוח יומי - מערכת ניהול מלאי</h1>
                    <p><strong>תאריך:</strong> {today.strftime('%d/%m/%Y')}</p>

                    <h2>📈 סטטיסטיקות כלליות</h2>
                    <div class="stats">
                        <div class="stat-card">
                            <div class="stat-value">{total_products}</div>
                            <div class="stat-label">סה"כ מוצרים</div>
                        </div>
                        <div class="stat-card success">
                            <div class="stat-value">₪{total_stock_value:,.0f}</div>
                            <div class="stat-label">ערך מלאי</div>
                        </div>
                        <div class="stat-card warning">
                            <div class="stat-value">{low_stock}</div>
                            <div class="stat-label">מלאי נמוך</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">{daily_sales_count}</div>
                            <div class="stat-label">מכירות היום</div>
                        </div>
                    </div>

                    <h2>💰 מכירות היום</h2>
                    <div class="stats">
                        <div class="stat-card success">
                            <div class="stat-value">{daily_sales_count}</div>
                            <div class="stat-label">כמות מכירות</div>
                        </div>
                        <div class="stat-card">
                            <div class="stat-value">₪{daily_sales_amount:,.2f}</div>
                            <div class="stat-label">סה"כ מכירות</div>
                        </div>
                    </div>
            """

            # התראות קריטיות
            if critical_alerts > 0 or out_of_stock > 0:
                html_content += "<h2>⚠️ התראות חשובות</h2>"
                if critical_alerts > 0:
                    html_content += f'<div class="alert alert-danger"><strong>🚨 {critical_alerts} התראות קריטיות!</strong></div>'
                if out_of_stock > 0:
                    html_content += f'<div class="alert alert-danger"><strong>📦 {out_of_stock} מוצרים אזלו מהמלאי!</strong></div>'

            # מוצרים במלאי נמוך
            if low_stock_products:
                html_content += """
                    <h2>📉 מוצרים במלאי נמוך (TOP 10)</h2>
                    <table>
                        <thead>
                            <tr>
                                <th>מוצר</th>
                                <th>כמות במלאי</th>
                                <th>מינימום</th>
                                <th>סטטוס</th>
                            </tr>
                        </thead>
                        <tbody>
                """
                for product in low_stock_products:
                    status = "🔴 אזל" if product.quantity == 0 else "🟡 נמוך"
                    html_content += f"""
                            <tr>
                                <td><strong>{product.name}</strong></td>
                                <td>{product.quantity}</td>
                                <td>{product.min_quantity}</td>
                                <td>{status}</td>
                            </tr>
                    """
                html_content += """
                        </tbody>
                    </table>
                """

            html_content += """
                    <div class="alert" style="background-color: #d1ecf1; border-right: 4px solid #0c5460; color: #0c5460; margin: 20px 0;">
                        <strong>📎 קבצים מצורפים:</strong><br>
                        קובץ ZIP עם כל הדוחות המפורטים (מוצרים, מכירות, לקוחות, קטגוריות, ספקים, מיקומים, התראות + מלאי נמוך)
                    </div>

                    <div class="footer">
                        <p>דוח מיידי נוצר לבקשתך ממערכת ניהול המלאי</p>
                        <p>© 2025 מערכת ניהול מלאי</p>
                    </div>
                </div>
            </body>
            </html>
            """

            # יצירת קובץ ZIP עם כל הדוחות
            import zipfile
            zip_buffer = io.BytesIO()

            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # 1. מוצרים
                products_buffer = io.BytesIO()
                products_buffer.write('\ufeff'.encode('utf-8'))
                writer = csv.writer(io.TextIOWrapper(products_buffer, encoding='utf-8', newline=''))
                writer.writerow(['שם', 'SKU', 'ברקוד', 'קטגוריה', 'כמות', 'מחיר מכירה', 'מחיר עלות', 'ערך מלאי', 'ספק'])
                products = Product.objects.select_related('category', 'supplier').all()
                for product in products:
                    writer.writerow([
                        product.name, product.sku or '', product.barcode or '',
                        product.category.name if product.category else '',
                        product.quantity, product.selling_price or 0, product.cost_price or 0,
                        product.stock_value, product.supplier.name if product.supplier else ''
                    ])
                zip_file.writestr('1_products.csv', products_buffer.getvalue())

                # 2. מכירות
                sales_buffer = io.BytesIO()
                sales_buffer.write('\ufeff'.encode('utf-8'))
                writer = csv.writer(io.TextIOWrapper(sales_buffer, encoding='utf-8', newline=''))
                writer.writerow(['מספר חשבונית', 'תאריך', 'לקוח', 'סכום כולל', 'סכום מס', 'מיקום', 'אמצעי תשלום', 'סטטוס'])
                sales = Sale.objects.select_related('customer', 'location').all()
                for sale in sales:
                    writer.writerow([
                        sale.invoice_number, sale.created_at.strftime('%d/%m/%Y %H:%M'),
                        sale.customer.name if sale.customer else '',
                        sale.total_amount, sale.tax_amount or 0,
                        sale.location.name if sale.location else '',
                        sale.get_payment_method_display() if sale.payment_method else '',
                        sale.get_status_display() if sale.status else ''
                    ])
                zip_file.writestr('2_sales.csv', sales_buffer.getvalue())

                # 3. לקוחות
                customers_buffer = io.BytesIO()
                customers_buffer.write('\ufeff'.encode('utf-8'))
                writer = csv.writer(io.TextIOWrapper(customers_buffer, encoding='utf-8', newline=''))
                writer.writerow(['שם', 'קוד לקוח', 'ח.פ/ע.מ', 'אימייל', 'טלפון', 'כתובת', 'הנחה %', 'מגבלת אשראי', 'סה"כ רכישות', 'פעיל'])
                customers = Customer.objects.all()
                for customer in customers:
                    writer.writerow([
                        customer.name, customer.customer_code or '', customer.tax_id or '',
                        customer.email or '', customer.phone or '', customer.address or '',
                        customer.discount_percent or 0, customer.credit_limit or 0,
                        customer.total_purchases or 0,
                        'כן' if customer.is_active else 'לא'
                    ])
                zip_file.writestr('3_customers.csv', customers_buffer.getvalue())

                # 4. קטגוריות
                categories_buffer = io.BytesIO()
                categories_buffer.write('\ufeff'.encode('utf-8'))
                writer = csv.writer(io.TextIOWrapper(categories_buffer, encoding='utf-8', newline=''))
                writer.writerow(['שם', 'תיאור', 'כמות מוצרים'])
                categories = Category.objects.annotate(product_count=Count('product')).all()
                for category in categories:
                    writer.writerow([
                        category.name, category.description or '', category.product_count
                    ])
                zip_file.writestr('4_categories.csv', categories_buffer.getvalue())

                # 5. ספקים
                suppliers_buffer = io.BytesIO()
                suppliers_buffer.write('\ufeff'.encode('utf-8'))
                writer = csv.writer(io.TextIOWrapper(suppliers_buffer, encoding='utf-8', newline=''))
                writer.writerow(['שם', 'איש קשר', 'אימייל', 'טלפון', 'כתובת', 'כמות מוצרים', 'פעיל'])
                suppliers = Supplier.objects.annotate(product_count=Count('product')).all()
                for supplier in suppliers:
                    writer.writerow([
                        supplier.name, supplier.contact_person or '', supplier.email or '',
                        supplier.phone or '', supplier.address or '', supplier.product_count,
                        'כן' if supplier.is_active else 'לא'
                    ])
                zip_file.writestr('5_suppliers.csv', suppliers_buffer.getvalue())

                # 6. מיקומים
                locations_buffer = io.BytesIO()
                locations_buffer.write('\ufeff'.encode('utf-8'))
                writer = csv.writer(io.TextIOWrapper(locations_buffer, encoding='utf-8', newline=''))
                writer.writerow(['שם', 'סוג', 'כתובת', 'מנהל', 'כמות מוצרים', 'פעיל'])
                locations = Location.objects.annotate(product_count=Count('product_stocks')).all()
                for location in locations:
                    writer.writerow([
                        location.name, location.get_location_type_display(), location.address or '',
                        location.manager_name or '', location.product_count,
                        'כן' if location.is_active else 'לא'
                    ])
                zip_file.writestr('6_locations.csv', locations_buffer.getvalue())

                # 7. התראות
                alerts_buffer = io.BytesIO()
                alerts_buffer.write('\ufeff'.encode('utf-8'))
                writer = csv.writer(io.TextIOWrapper(alerts_buffer, encoding='utf-8', newline=''))
                writer.writerow(['מוצר', 'סוג התראה', 'עדיפות', 'הודעה', 'נפתר', 'תאריך יצירה'])
                alerts = Alert.objects.select_related('product').all()
                for alert in alerts:
                    writer.writerow([
                        alert.product.name if alert.product else '',
                        alert.get_alert_type_display(), alert.get_priority_display(),
                        alert.message or '', 'כן' if alert.is_resolved else 'לא',
                        alert.created_at.strftime('%d/%m/%Y %H:%M') if alert.created_at else ''
                    ])
                zip_file.writestr('7_alerts.csv', alerts_buffer.getvalue())

                # 8. מלאי נמוך (דוח מיוחד)
                if low_stock_products:
                    low_stock_buffer = io.BytesIO()
                    low_stock_buffer.write('\ufeff'.encode('utf-8'))
                    writer = csv.writer(io.TextIOWrapper(low_stock_buffer, encoding='utf-8', newline=''))
                    writer.writerow(['שם מוצר', 'SKU', 'כמות במלאי', 'מינימום', 'קטגוריה', 'סטטוס'])
                    for product in low_stock_products:
                        status = '🔴 אזל' if product.quantity == 0 else '🟡 נמוך'
                        writer.writerow([
                            product.name, product.sku or '', product.quantity,
                            product.min_quantity, product.category.name if product.category else '',
                            status
                        ])
                    zip_file.writestr('8_low_stock_urgent.csv', low_stock_buffer.getvalue())

            # הכנת הקובץ לצירוף
            zip_buffer.seek(0)
            attachments = [(f'reports_{today.strftime("%Y%m%d")}.zip', zip_buffer.getvalue(), 'application/zip')]

            # שליחת המייל
            from django.core.mail import EmailMessage
            email = EmailMessage(
                subject=f'⚡ דוח מיידי - {today.strftime("%d/%m/%Y")} {timezone.now().strftime("%H:%M")}',
                body=html_content,
                from_email=settings_obj.default_from_email or settings_obj.email_host_user,
                to=[settings_obj.daily_report_email],
            )
            email.content_subtype = 'html'

            # הוספת קבצים מצורפים
            for filename, content, mimetype in attachments:
                email.attach(filename, content, mimetype)

            # שליחת המייל עם טיפול ב-SSL
            import ssl
            import smtplib
            try:
                email.send()
            except Exception as ssl_error:
                # נסה שוב עם אימות SSL מושבת
                if 'CERTIFICATE_VERIFY_FAILED' in str(ssl_error):
                    # צור קונטקסט SSL שמתעלם מאימות תעודות
                    context = ssl.create_default_context()
                    context.check_hostname = False
                    context.verify_mode = ssl.CERT_NONE
                    
                    # שלח ידנית עם ההגדרות המותאמות
                    connection = smtplib.SMTP(settings_obj.email_host, settings_obj.email_port)
                    if settings_obj.email_use_tls:
                        connection.starttls(context=context)
                    connection.login(settings_obj.email_host_user, settings_obj.email_host_password)
                    connection.send_message(email.message())
                    connection.quit()
                else:
                    raise ssl_error

            return JsonResponse({
                'success': True,
                'message': f'✅ הדוח נשלח בהצלחה ל-{settings_obj.daily_report_email}!',
                'details': {
                    'products': total_products,
                    'sales': daily_sales_count,
                    'alerts': critical_alerts
                }
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'❌ שגיאה בשליחת הדוח: {str(e)}'
            })

    return JsonResponse({'success': False, 'error': 'Method not allowed'})
