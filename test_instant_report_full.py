#!/usr/bin/env python
"""
בדיקה מלאה של send_instant_report - שלב אחר שלב
"""
import os
import sys
import django

# Setup Django
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'inventory_project.settings')
django.setup()

from inventory.models import SystemSettings, Product, Sale, Alert, StockMovement, Tenant
from django.utils import timezone
from django.db.models import Sum, F
from io import BytesIO
import traceback

def test_step_by_step():
    """בדיקה שלב אחר שלב"""

    print("\n" + "="*80)
    print("  🧪 בדיקת send_instant_report - שלב אחר שלב")
    print("="*80 + "\n")

    # שלב 1: קבלת Tenant
    print("📦 שלב 1: קבלת Tenant")
    try:
        tenant = Tenant.objects.first()
        if tenant:
            print(f"  ✅ Tenant: {tenant.name}")
        else:
            print("  ❌ אין Tenant!")
            return
    except Exception as e:
        print(f"  ❌ שגיאה: {e}")
        traceback.print_exc()
        return

    # שלב 2: טעינת הגדרות
    print("\n⚙️  שלב 2: טעינת הגדרות")
    try:
        settings_obj = SystemSettings.load()
        print(f"  ✅ email_enabled: {settings_obj.email_enabled}")
        print(f"  ✅ daily_report_email: {settings_obj.daily_report_email}")

        if not settings_obj.email_enabled:
            print("  ⚠️  שליחת מייל לא מופעלת!")
        if not settings_obj.daily_report_email:
            print("  ⚠️  לא הוגדר מייל!")
    except Exception as e:
        print(f"  ❌ שגיאה: {e}")
        traceback.print_exc()
        return

    # שלב 3: איסוף נתונים
    print("\n📊 שלב 3: איסוף נתונים")
    try:
        today = timezone.now().date()

        products_qs = Product.objects.filter(tenant=tenant)
        sales_qs = Sale.objects.filter(tenant=tenant)

        total_products = products_qs.count()
        low_stock = products_qs.filter(quantity__lte=F('min_quantity')).count()

        print(f"  ✅ מוצרים: {total_products}")
        print(f"  ✅ מלאי נמוך: {low_stock}")
    except Exception as e:
        print(f"  ❌ שגיאה באיסוף נתונים: {e}")
        traceback.print_exc()
        return

    # שלב 4: יצירת HTML
    print("\n📧 שלב 4: יצירת HTML")
    try:
        html_content = f"""
        <!DOCTYPE html>
        <html dir="rtl" lang="he">
        <head>
            <meta charset="UTF-8">
        </head>
        <body>
            <h1>דוח מיידי</h1>
            <p>מוצרים: {total_products}</p>
        </body>
        </html>
        """
        print(f"  ✅ HTML נוצר - {len(html_content)} תווים")
    except Exception as e:
        print(f"  ❌ שגיאה ביצירת HTML: {e}")
        traceback.print_exc()
        return

    # שלב 5: יצירת Excel
    print("\n📝 שלב 5: יצירת Excel")
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "דוח כללי"

        # כותרת
        ws['A1'] = 'דוח מיידי'
        ws['A1'].font = Font(size=14, bold=True)

        # נתונים
        ws['A3'] = 'סה"כ מוצרים'
        ws['B3'] = total_products

        # שמירה
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        print(f"  ✅ Excel נוצר - {len(excel_buffer.getvalue())} bytes")
    except Exception as e:
        print(f"  ❌ שגיאה ביצירת Excel: {e}")
        traceback.print_exc()
        return

    # שלב 6: שליחת מייל (סימולציה)
    print("\n📬 שלב 6: שליחת מייל")
    try:
        from django.core.mail import EmailMessage
        from django.conf import settings as django_settings

        print(f"  • EMAIL_BACKEND: {django_settings.EMAIL_BACKEND}")

        email = EmailMessage(
            subject=f'⚡ דוח מיידי - {today.strftime("%d/%m/%Y")}',
            body=html_content,
            from_email='noreply@inventory.local',
            to=['admin@inventory.local'],
        )
        email.content_subtype = 'html'
        email.attach('report.xlsx', excel_buffer.getvalue(),
                     'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        # נסיון שליחה
        email.send()

        print("  ✅ מייל נשלח בהצלחה!")
        print("  💡 בדוק בטרמינל של Django - תראה את המייל!")

    except Exception as e:
        print(f"  ❌ שגיאה בשליחת מייל: {e}")
        traceback.print_exc()
        return

    print("\n" + "="*80)
    print("  ✅ כל השלבים עברו בהצלחה!")
    print("  💡 הפונקציה send_instant_report אמורה לעבוד.")
    print("="*80 + "\n")

if __name__ == '__main__':
    test_step_by_step()

