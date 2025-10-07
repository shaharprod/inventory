#!/usr/bin/env python
"""
בדיקת שליחת דוח ישירות - סימולציה מלאה
"""
import os
import sys
import django

# Setup Django
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'inventory_project.settings')
django.setup()

from inventory.models import SystemSettings, Product, Sale, Alert, StockMovement, Tenant
from django.utils import timezone
from django.db.models import Sum, F
from io import BytesIO

def test_send_report():
    """סימולציה מלאה של send_instant_report"""

    print("\n" + "="*80)
    print("  🧪 בודק שליחת דוח - סימולציה מלאה")
    print("="*80 + "\n")

    try:
        # קבל tenant
        tenant = Tenant.objects.first()
        print(f"✅ Tenant: {tenant.name}\n")

        settings_obj = SystemSettings.load()
        print(f"✅ הגדרות נטענו\n")

        # איסוף נתונים
        today = timezone.now().date()

        # סינון לפי tenant
        products_qs = Product.objects.all()
        sales_qs = Sale.objects.all()
        alerts_qs = Alert.objects.all()
        movements_qs = StockMovement.objects.all()

        if tenant:
            products_qs = products_qs.filter(tenant=tenant)
            sales_qs = sales_qs.filter(tenant=tenant)

        print("📊 איסוף נתונים:")
        total_products = products_qs.count()
        print(f"  • מוצרים: {total_products}")

        low_stock = products_qs.filter(quantity__lte=F('min_quantity')).count()
        print(f"  • מלאי נמוך: {low_stock}")

        # בדיקת יצירת Excel - זה בדרך כלל הגורם לשגיאה
        print("\n📝 בודק יצירת Excel...")
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "דוח כללי"

            # כותרת
            ws['A1'] = 'דוח מיידי'
            ws['A2'] = f'תאריך: {today.strftime("%d/%m/%Y")}'

            # נתונים
            ws['A4'] = 'סטטיסטיקות'
            ws['A5'] = 'סה"כ מוצרים'
            ws['B5'] = total_products

            # שמירה ל-buffer
            excel_buffer = BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            print(f"  ✅ Excel נוצר בהצלחה! גודל: {len(excel_buffer.getvalue())} bytes")

        except Exception as excel_error:
            print(f"  ❌ שגיאה ביצירת Excel: {excel_error}")
            import traceback
            traceback.print_exc()
            return

        # בדיקת יצירת HTML
        print("\n📧 בודק יצירת HTML...")
        try:
            html_content = f"""
            <!DOCTYPE html>
            <html dir="rtl" lang="he">
            <head>
                <meta charset="UTF-8">
            </head>
            <body>
                <h1>דוח מיידי</h1>
                <p>סה"כ מוצרים: {total_products}</p>
            </body>
            </html>
            """
            print(f"  ✅ HTML נוצר בהצלחה! אורך: {len(html_content)} תווים")

        except Exception as html_error:
            print(f"  ❌ שגיאה ביצירת HTML: {html_error}")
            return

        print("\n✅ כל השלבים עברו בהצלחה!")
        print("\n💡 הבעיה כנראה בשליחת המייל עצמו או ב-middleware")
        print("   נסה להריץ בדיקה זו:")
        print("   1. בדוק את הקונסול של Django (חלון השרת)")
        print("   2. חפש שגיאות ב-Traceback")
        print("   3. העתק את השגיאה המלאה")

    except Exception as e:
        print(f"\n❌ שגיאה כללית: {e}\n")
        import traceback
        traceback.print_exc()

    print("\n" + "="*80 + "\n")

if __name__ == '__main__':
    test_send_report()

